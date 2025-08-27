import os
import cv2
from ultralytics import YOLO
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from PIL import Image as PILImage

# Rutas y parámetros
CARPETA_ORIGEN = "./Imagenes"
CARPETA_RECORTES = "./recortes"
RUTA_MODELO = "./best.pt"
RUTA_EXCEL_PROCESADOS = "./resultados/dorsales_recortes.xlsx"
RUTA_EXCEL_NO_DETECTADOS = "./resultados/no_detectados.xlsx"


PADDING_REL = 0.03  # porcentaje para ampliar la caja de detección

# Crear carpeta recortes si no existe
os.makedirs(CARPETA_RECORTES, exist_ok=True)

print("Cargando modelo YOLO...")
model = YOLO(RUTA_MODELO)

# --- Cargar o crear Excel PROCESADOS ---
if os.path.exists(RUTA_EXCEL_PROCESADOS):
    print(f"Cargando archivo existente {RUTA_EXCEL_PROCESADOS}")
    wb_procesados = load_workbook(RUTA_EXCEL_PROCESADOS)
    ws_procesados = wb_procesados.active
else:
    print("Creando nuevo archivo Excel para procesados")
    wb_procesados = Workbook()
    ws_procesados = wb_procesados.active
    ws_procesados.title = "Recortes"
    ws_procesados.cell(row=1, column=1, value="Archivo")
    ws_procesados.cell(row=1, column=2, value="Recorte Dorsal")
    ws_procesados.row_dimensions[1].height = 45
    ws_procesados.column_dimensions[get_column_letter(2)].width = 20

# Leer nombres ya procesados para evitar duplicados
procesados = set()
for row in ws_procesados.iter_rows(min_row=2, max_col=1):
    val = row[0].value
    if val:
        procesados.add(val)

print(f"Imágenes ya procesadas en Excel: {len(procesados)}")

# --- Cargar o crear Excel NO DETECTADOS ---
if os.path.exists(RUTA_EXCEL_NO_DETECTADOS):
    print(f"Cargando archivo existente {RUTA_EXCEL_NO_DETECTADOS}")
    wb_no_det = load_workbook(RUTA_EXCEL_NO_DETECTADOS)
    ws_no_det = wb_no_det.active

    # Leer no detectados existentes
    no_detectados_existentes = set()
    for row in ws_no_det.iter_rows(min_row=2, min_col=1, max_col=1):
        val = row[0].value
        if val:
            no_detectados_existentes.add(val)

    # Encontrar la última fila usada
    fila_no_det = ws_no_det.max_row + 1

else:
    print("Creando nuevo archivo Excel para no detectados")
    wb_no_det = Workbook()
    ws_no_det = wb_no_det.active
    ws_no_det.title = "No Detectados"
    ws_no_det.cell(row=1, column=1, value="Archivo")
    ws_no_det.cell(row=1, column=2, value="Enlace Foto")
    ws_no_det.row_dimensions[1].height = 45
    ws_no_det.column_dimensions[get_column_letter(1)].width = 40
    ws_no_det.column_dimensions[get_column_letter(2)].width = 60
    bold_font = Font(bold=True)
    ws_no_det["A1"].font = bold_font
    ws_no_det["B1"].font = bold_font

    no_detectados_existentes = set()
    fila_no_det = 2

lista_no_detectados = []

# Buscar imágenes en carpeta
archivos = [f for f in os.listdir(CARPETA_ORIGEN) if f.lower().endswith(('.jpg','.jpeg','.png'))]
print(f"Total archivos en carpeta: {len(archivos)}")

# Empezar desde la siguiente fila libre en procesados
row_procesados = ws_procesados.max_row + 1

for archivo in archivos:
    if archivo in procesados:
        print(f"Ya procesado, saltando: {archivo}")
        continue

    ruta_img = os.path.join(CARPETA_ORIGEN, archivo)
    img = cv2.imread(ruta_img)
    if img is None:
        print(f"No se pudo abrir {archivo}")
        lista_no_detectados.append(archivo)
        continue

    resultados = model(img)

    if len(resultados) == 0 or len(resultados[0].boxes) == 0:
        print(f"No se detectó dorsal en {archivo}")
        lista_no_detectados.append(archivo)
        continue

    boxes = resultados[0].boxes
    box_conf = [(b.conf[0].item(), b) for b in boxes]
    box_conf.sort(key=lambda x: x[0], reverse=True)
    mejor_box = box_conf[0][1]

    x1, y1, x2, y2 = map(int, mejor_box.xyxy[0])

    ancho = x2 - x1
    alto = y2 - y1
    pad_x = int(ancho * PADDING_REL)
    pad_y = int(alto * PADDING_REL)

    x1n = max(0, x1 - pad_x)
    y1n = max(0, y1 - pad_y)
    x2n = min(img.shape[1], x2 + pad_x)
    y2n = min(img.shape[0], y2 + pad_y)

    recorte = img[y1n:y2n, x1n:x2n]

    nombre_recorte = os.path.splitext(archivo)[0] + "_recorte.jpg"
    ruta_recorte = os.path.join(CARPETA_RECORTES, nombre_recorte)
    cv2.imwrite(ruta_recorte, recorte)

    pil_img = PILImage.open(ruta_recorte)
    ancho_original, alto_original = pil_img.size

    ancho_nuevo = 100
    alto_nuevo = int(alto_original * ancho_nuevo / ancho_original)

    img_xl = XLImage(ruta_recorte)
    img_xl.width = ancho_nuevo
    img_xl.height = alto_nuevo

    ws_procesados.cell(row=row_procesados, column=1, value=archivo)
    celda_imagen = f"B{row_procesados}"
    ws_procesados.add_image(img_xl, celda_imagen)
    ws_procesados.row_dimensions[row_procesados].height = 45

    print(f"Procesado {archivo}")
    row_procesados += 1

# Añadir no detectados solo si no están ya en el Excel
for archivo_no_det in lista_no_detectados:
    if archivo_no_det not in no_detectados_existentes:
        ws_no_det.cell(row=fila_no_det, column=1, value=archivo_no_det)

        ruta_abs = os.path.abspath(os.path.join(CARPETA_ORIGEN, archivo_no_det))
        url_file = f"file:///{ruta_abs.replace(os.sep, '/')}"

        ws_no_det.cell(row=fila_no_det, column=2, value="Abrir imagen")
        ws_no_det.cell(row=fila_no_det, column=2).hyperlink = url_file
        ws_no_det.cell(row=fila_no_det, column=2).font = Font(color="0000FF", underline="single")

        ws_no_det.row_dimensions[fila_no_det].height = 30
        fila_no_det += 1

wb_procesados.save(RUTA_EXCEL_PROCESADOS)
print(f"Archivo Excel de procesados guardado en {RUTA_EXCEL_PROCESADOS}")

wb_no_det.save(RUTA_EXCEL_NO_DETECTADOS)
print(f"Archivo Excel de no detectados guardado en {RUTA_EXCEL_NO_DETECTADOS}")
